import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
import json
import os
from typing import Optional

from . import head_type
from .exceptions import (
    TableFileNotFoundError,
    InvalidFileFormatError,
    EmptyFileError,
    MissingHeaderMarkerError,
    MissingDataMarkerError,
)


class Head:
    def __init__(
        self, sheet: Worksheet, row: tuple[Cell, ...], filename: Optional[str] = None
    ) -> None:
        self.sheet = sheet
        self.column = len(row)
        self.filename = filename
        self.head = head_type.head_creator(row[0], filename)
        self.headList: list[head_type.HeadType] = [self.head] * self.column

    def get_cell_max_col(self, cell: Cell) -> int:
        for i in self.sheet.merged_cells.ranges:
            if cell.coordinate in i:
                return i.max_col - 1
        return cell.column - 1

    def row_parser(self, row: tuple[Cell, ...]) -> None:
        i = 0
        while i < self.column:
            if row[i].value:
                h = head_type.head_creator(row[i], self.filename)
                j = self.get_cell_max_col(row[i])
                self.headList[i].add_child(h)
                self.headList[i:j] = [h] * (j - i)
                i = j
            else:
                i += 1


def get_data(filename: str) -> head_type.data:
    # Validate input
    if not filename:
        raise ValueError("Filename cannot be empty")

    if not os.path.exists(filename):
        raise TableFileNotFoundError(filename)

    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise InvalidFileFormatError(filename)

    # Load workbook
    try:
        workbook = openpyxl.load_workbook(filename)
    except Exception as e:
        raise InvalidFileFormatError(filename) from e

    if not workbook.active:
        raise EmptyFileError(filename)

    iter_rows = workbook.active.iter_rows()

    # Check first row
    try:
        first_row = next(iter_rows)
    except StopIteration:
        raise EmptyFileError(filename)

    # Validate #head marker
    first_cell = first_row[0]
    if first_cell.value != "#head":
        raise MissingHeaderMarkerError(
            first_cell, str(first_cell.value) if first_cell.value else "empty", filename
        )

    # Parse header
    head = Head(workbook.active, first_row[1:], filename)

    # Parse rows
    is_data = False
    data_root = None

    for row in iter_rows:
        if row[0].value == "#":
            continue
        elif row[0].value == "#data":
            is_data = True
            data_root = head.head.parse_data(row[1:], True, filename)
            continue

        if is_data:
            head.head.parse_data(row[1:], False, filename)
        else:
            head.row_parser(row[1:])

    # Validate data was found
    if data_root is None:
        raise MissingDataMarkerError(filename)

    return data_root


def get_json(filename: str, indent: int = 2) -> str:
    """Convert Excel file to JSON string

    Args:
        filename: Path to Excel file
        indent: JSON indentation level (default 2)

    Returns:
        Formatted JSON string
    """
    return json.dumps(get_data(filename), indent=indent)


if __name__ == "__main__":
    print(get_json("Table.xlsx"))
