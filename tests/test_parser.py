import unittest
import json
import os
import openpyxl
from src import stdem
from src.stdem.TableException import (
    FileNotFoundError as TableFileNotFoundError,
    InvalidFileFormatError,
    EmptyFileError,
    MissingHeaderMarkerError,
    MissingDataMarkerError,
    InvalidTypeNameError,
    InvalidHeaderFormatError
)


class TestParser(unittest.TestCase):

    def test_example_excel(self):
        """Test parsing example.xlsx and compare with expected JSON"""
        result = stdem.ExcelParser.getData("tests/excel/example.xlsx")

        with open("tests/json/example.json", "r", encoding="utf-8") as f:
            expected = json.load(f)

        self.assertEqual(result, expected)

    def test_unit_data_excel(self):
        """Test parsing UnitData.xlsx and compare with expected JSON"""
        result = stdem.ExcelParser.getData("tests/excel/UnitData.xlsx")

        with open("tests/json/UnitData.json", "r", encoding="utf-8") as f:
            expected = json.load(f)

        self.assertEqual(result, expected)

    def test_skill_table_excel(self):
        """Test parsing SkillTable.xlsx and compare with expected JSON"""
        result = stdem.ExcelParser.getData("tests/excel/SkillTable.xlsx")

        with open("tests/json/SkillTable.json", "r", encoding="utf-8") as f:
            expected = json.load(f)

        self.assertEqual(result, expected)

    def test_effect_table_excel(self):
        """Test parsing EffectTable.xlsx and compare with expected JSON"""
        result = stdem.ExcelParser.getData("tests/excel/EffectTable.xlsx")

        with open("tests/json/EffectTable.json", "r", encoding="utf-8") as f:
            expected = json.load(f)

        self.assertEqual(result, expected)

    def test_getJson_returns_formatted_json(self):
        """Test that getJson returns formatted JSON with indentation"""
        json_str = stdem.ExcelParser.getJson("tests/excel/example.xlsx")

        # Should be valid JSON
        parsed = json.loads(json_str)
        self.assertIsInstance(parsed, dict)

        # Should have indentation
        self.assertIn("\n", json_str)

        # Compare with expected
        with open("tests/json/example.json", "r", encoding="utf-8") as f:
            expected = json.load(f)
        self.assertEqual(parsed, expected)


class TestErrorHandling(unittest.TestCase):
    """Test error handling for various invalid inputs"""

    def test_empty_filename(self):
        """Test that empty filename raises ValueError"""
        with self.assertRaises(ValueError) as context:
            stdem.ExcelParser.getData("")
        self.assertIn("cannot be empty", str(context.exception))

    def test_nonexistent_file(self):
        """Test that nonexistent file raises FileNotFoundError"""
        with self.assertRaises(TableFileNotFoundError) as context:
            stdem.ExcelParser.getData("tests/excel/nonexistent.xlsx")
        self.assertIn("not found", str(context.exception))

    def test_invalid_file_format(self):
        """Test that non-xlsx file raises InvalidFileFormatError"""
        with self.assertRaises(InvalidFileFormatError):
            stdem.ExcelParser.getData("tests/test_parser.py")

    def test_missing_head_marker(self):
        """Test that file without #head marker raises error"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'invalid'
        ws['B1'] = 'name:string'
        test_file = "tests/excel/test_no_head.xlsx"

        try:
            wb.save(test_file)
            with self.assertRaises(MissingHeaderMarkerError) as context:
                stdem.ExcelParser.getData(test_file)
            self.assertIn("#head", str(context.exception))
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_missing_data_marker(self):
        """Test that file without #data marker raises error"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '#head'
        ws['B1'] = 'name:string'
        test_file = "tests/excel/test_no_data.xlsx"

        try:
            wb.save(test_file)
            with self.assertRaises(MissingDataMarkerError) as context:
                stdem.ExcelParser.getData(test_file)
            self.assertIn("#data", str(context.exception))
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_invalid_type_name(self):
        """Test that invalid type name raises error"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '#head'
        ws['B1'] = 'name:invalid_type'
        test_file = "tests/excel/test_invalid_type.xlsx"

        try:
            wb.save(test_file)
            with self.assertRaises(InvalidTypeNameError) as context:
                stdem.ExcelParser.getData(test_file)
            self.assertIn("Invalid type", str(context.exception))
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)

    def test_invalid_header_format(self):
        """Test that header without colon raises error"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '#head'
        ws['B1'] = 'namestring'  # Missing colon
        test_file = "tests/excel/test_invalid_format.xlsx"

        try:
            wb.save(test_file)
            with self.assertRaises(InvalidHeaderFormatError) as context:
                stdem.ExcelParser.getData(test_file)
            self.assertIn("format", str(context.exception).lower())
        finally:
            if os.path.exists(test_file):
                os.remove(test_file)


if __name__ == "__main__":
    unittest.main()
